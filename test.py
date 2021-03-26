import json
import sys
import random
import xlrd
import xlwt
sys.path.append(sys.path[0] + "/bert-utils-master/")
import extract_feature
import numpy as np
import time
from scipy import spatial
import heapq
import jieba
import re
import openpyxl
import math
import urllib.request
import requests
import traceback

class BM25(object):

    def __init__(self, docs):
        self.D = len(docs)
        self.avgdl = sum([len(doc)+0.0 for doc in docs]) / (self.D + 1)
        self.docs = docs
        self.f = []  # 列表的每一个元素是一个dict，dict存储着一个文档中每个词的出现次数
        self.df = {} # 存储每个词及出现了该词的文档数量
        self.idf = {} # 存储每个词的idf值
        self.k1 = 1.5
        self.b = 0.75
        self.init()

    def init(self):
        for doc in self.docs:
            tmp = {}
            for word in doc:
                tmp[word] = tmp.get(word, 0) + 1  # 存储每个文档中每个词的出现次数
            self.f.append(tmp)
            for k in tmp.keys():
                self.df[k] = self.df.get(k, 0) + 1
        for k, v in self.df.items():
            self.idf[k] = math.log(self.D-v+0.5)-math.log(v+0.5)

    def sim(self, doc, index):
        score = 0
        for word in doc:
            if word not in self.f[index]:
                continue
            d = len(self.docs[index])
            score += (self.idf[word]*self.f[index][word]*(self.k1+1)
                      / (self.f[index][word]+self.k1*(1-self.b+self.b*d
                                                      / self.avgdl)))
        return score

    def simall(self, doc):
        scores = []
        for index in range(self.D):
            score = self.sim(doc, index)
            scores.append(score)
        return scores

def get_query_description(query,text):#获得title下关于query的description
    #print(query,text)
    max_length = 100
    s = BM25(text)
    score = s.simall(query)
    result_score = {}
    for i in range(len(text)):
        result_score[text[i]] = score[i]
    a = sorted(result_score.items(), key=lambda x: x[1], reverse=True)
    result_str = ""
    str_index = text.index(a[0][0])
    result_str = "".join(text[str_index:])
    high_line = []
    if len(result_str) < max_length:
        for word in query:
            if word in result_str:
                high_line.append(word)
        return result_str,high_line
    else:
        for word in query:
            if word in result_str[:max_length]:
                high_line.append(word)
        return result_str[:max_length],high_line


def data_process():
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/百科问答标注语料_1203.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    
    intent2data = {}
    for i in range(1,rows):
        row_data = table.row_values(i)
        if row_data[1] != "":
            if row_data[1] not in intent2data:
                intent2data[row_data[1]] = [row_data[0]]
            else:
                if len(intent2data[row_data[1]]) < 3:
                    intent2data[row_data[1]].append(row_data[0])
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/百科问答意图体系1203.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    # 创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')
    k = 1
    for i in range(1,rows):
        row_data = table.row_values(i)
        for j in range(3):
            worksheet.write(k, j, label = row_data[j])
        if row_data[2] != "":
            intent = row_data[2]
        else:
            intent = row_data[1]
        if intent in intent2data:
            if row_data[3] != "":
                worksheet.write(k, 3, label = row_data[3])
                for j in range(4,min(6,4 + len(intent2data[intent]))):
                    worksheet.write(k, j, label = intent2data[intent][j - 4])
            else:  
                for j in range(3,3 + len(intent2data[intent])):
                    worksheet.write(k, j, label = intent2data[intent][j - 3])
        worksheet.write(k, 6, label = row_data[4])
        k += 1
    workbook.save('stastic/webSearch_data/百科问答标注语料_1208.xlsx')
    '''
    random.shuffle(table_data)

    # 创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')

    k = 1
    for i in range(1,int(len(table_data) * 0.9)):
        row_data = table_data[i]
        for j in range(len(row_data)):
            worksheet.write(k, j, label = row_data[j])
        k += 1
    workbook.save('stastic/webSearch_data/CQA_data_train.xlsx')

    # 创建一个workbook 设置编码
    workbook1 = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet1 = workbook1.add_sheet('sheet 1')

    k = 1
    for i in range(int(len(table_data) * 0.9),len(table_data)):
        row_data = table_data[i]
        for j in range(len(row_data)):
            worksheet1.write(k, j, label = row_data[j])
        k += 1
    workbook1.save('stastic/webSearch_data/CQA_data_test.xlsx')
    '''

def cosine_similarity(x, y, norm=False):
    
    vector_a = np.mat(x)
    vector_b = np.mat(y)
    num = float(vector_a * vector_b.T)
    denom = np.linalg.norm(vector_a) * np.linalg.norm(vector_b)
    cos = num / denom
    sim = 0.5 + 0.5 * cos
    return sim

bertvector_model = None

def get_bertvector():
    global bertvector_model
    if bertvector_model is None:
        bertvector_model = extract_feature.BertVector()
        print("load bertvector_model successfull!")
    return bertvector_model

query2encode = {}

def get_query2encode():
    global query2encode
    if query2encode == {}:
        bert = get_bertvector()
        data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/query_data_4w5_processed.xlsx')
        table = data.sheet_by_index(0)
        rows = table.nrows
        for i in range(1,rows):
            if i % 100 == 0:
                print(i)
            try:
                row_data = table.row_values(i)
                intent = row_data[0]
                if intent not in query2encode:
                    query2encode[intent] =  bert.encode([intent])[0].tolist() 
            except Exception as e:
                print(e)
        with open(sys.path[0] + "/stastic/webSearch_data/query2encode_4w5_processed_30_bert.json", "w", encoding='utf-8') as fp:
            fp.write(json.dumps(query2encode, ensure_ascii=False, indent=4))

        # data_list = []
        # data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/cross-validation/train_data_0.xlsx')
        # table = data.sheet_by_index(0)
        # rows = table.nrows
        # for i in range(1,rows):
        #     row_data = table.row_values(i)
        #     if row_data[0] not in data_list and row_data[0] != "":
        #         data_list.append(row_data[0])
        '''
        print("开始载入query2encode！")
        with open(sys.path[0] + '/stastic/webSearch_data/query2encode_3w7_30.json','r',encoding='utf8') as fp:
            json_data = json.load(fp)
        fp.close()
        for key in json_data:
            if key not in query2encode:
                query2encode[key] = np.array(json_data[key])

        with open(sys.path[0] + '/stastic/webSearch_data/query2encode_7k_30.json','r',encoding='utf8') as fp:
            json_data = json.load(fp)
        fp.close()
        for key in json_data:
            if key not in query2encode:
                query2encode[key] = np.array(json_data[key])
       
        with open(sys.path[0] + '/stastic/webSearch_data/query2encode_intent_30.json','r',encoding='utf8') as fp:
            json_data = json.load(fp)
        fp.close()
        for key in json_data:
            if key not in query2encode:
                query2encode[key] = np.array(json_data[key])
        with open(sys.path[0] + '/stastic/webSearch_data/query2encode_2W_30.json','r',encoding='utf8') as fp:
            json_data = json.load(fp)
        fp.close()
        for key in json_data:
            if key not in query2encode:
                query2encode[key] = np.array(json_data[key])
        print("载入query2encode成功！")
        '''
    return query2encode

cop = re.compile("[^\u4e00-\u9fa5^，。,]") # 匹配不是中文、大小写、数字的其他字符
def json_to_text(json_name):#json文件处理成文本
    with open(json_name,'r',encoding='utf8') as fp:
        json_data = json.load(fp)
        text_str = json.dumps(json_data,ensure_ascii=False)
        text_str = cop.sub('', text_str) #去除其他符号
    result_text = ""
    for text in text_str.split(","):
        if text != "":
            result_text = result_text + text + "，"
    #result_text.replace("。，","。")
    #print(result_text)
    return result_text

doc = {}
def get_doc():
    global doc
    if doc == {}:
        with open(sys.path[0] + '/stastic/webSearch_data/all_data.json','r',encoding='utf8') as fp:
            json_data = json.load(fp)
        fp.close()
        with open(sys.path[0] + '/stastic/webSearch_data/UrlFile_map.json','r',encoding='utf8') as fp:
            UrlFile_map_data = json.load(fp)
        fp.close()
        aircor_list = ["海南航空" , "南方航空", "中国国际航空"]
        for aircor in aircor_list:
            for kk in json_data[aircor]:
                for key in json_data[aircor][kk]:
                    for sub_dict in json_data[aircor][kk][key]:
                        for sub_key in sub_dict:
                            sub_key_text = ""
                            try:
                                if sub_dict[sub_key] in UrlFile_map_data:
                                    sub_dict[sub_key] = UrlFile_map_data[sub_dict[sub_key]]
                                sub_key_text = json_to_text(sys.path[0] + '/stastic/webSearch_data/json_data/' + sub_dict[sub_key])
                                if sub_key_text != "":
                                    doc[aircor + "@" + kk + "@" + key + "@" + sub_key] = jieba.lcut(sub_key_text,cut_all = False)
                            except Exception as e:
                                    pass
        print("载入doc成功！")
    return doc

intent2answer = {}
def get_intent2answer(): 
    global query2intent
    if intent2answer == {}:
        data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/schema_withanswer_1211.xlsx')
        table = data.sheet_by_index(0)
        rows = table.nrows
        for i in range(1,rows):
            row_data = table.row_values(i)
            if row_data[0] != "":
                intent = row_data[0]
            if row_data[1] != "":
                intent = row_data[1]
            if row_data[2] != "":
                intent = row_data[2]
            if intent not in intent2answer:
                if row_data[6] != "":
                    source = row_data[6]
                else:
                    source = "新增意图"
                answer = []
                for j in range(7,len(row_data)):
                    if row_data[j] != "":
                        answer.append(row_data[j])
                intent2answer[intent] = {"answer":answer,"source":source,"intent_level_1":row_data[0],"intent_level_2":row_data[1],"intent_level_3":row_data[2]}
        print("载入intent2answer成功！")
    return intent2answer

query2intent = {}
def get_query2intent():
    global query2intent
    if query2intent == {}:
        data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/final_data_4w2.xlsx')
        table = data.sheet_by_index(0)
        rows = table.nrows
        for i in range(1,rows):
            row_data = table.row_values(i)
            if row_data[0] not in query2intent and row_data[1] != "":
                query2intent[row_data[0]] = row_data[1].split(";")
        
        # data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/final_data_7k.xlsx')
        # table = data.sheet_by_index(0)
        # rows = table.nrows
        # for i in range(1,rows):
        #     row_data = table.row_values(i)
        #     if row_data[0] not in query2intent and row_data[1] != "":
        #         query2intent[row_data[0]] = [row_data[1]]
        data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/final_data_2w.xlsx')
        table = data.sheet_by_index(0)
        rows = table.nrows
        for i in range(1,rows):
            row_data = table.row_values(i)
            if row_data[0] not in query2intent and row_data[1] != "":
                query2intent[row_data[0]] = [row_data[1]]

        data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/schema_withanswer_1211.xlsx')
        table = data.sheet_by_index(0)
        rows = table.nrows
        for i in range(1,rows):
            row_data = table.row_values(i)
            if "KBQA" in row_data[6]:
                continue
            if row_data[0] != "" :
                intent = row_data[0]
            if row_data[1] != "" :
                intent = row_data[1]
            if row_data[2] != "" :
                intent = row_data[2]
            if intent not in query2intent:
                query2intent[intent] = [intent]

        print("载入query2intent成功！")
    return query2intent


def check_aircor(sent):#航空公司
    aircor_dict = {'南方航空':'南航','海南航空':'海航','中国国际航空':'国航'}
    for key in aircor_dict:
        if key in sent or aircor_dict[key] in sent:
            return key
    aircor_code = check_aircor_code(sent)
    if aircor_code == "00":
        return "所有航司"
    else:
        return "其他航司"

def check_aircor_code(sent):#获得航司代码
    file = open(sys.path[0] + "/stastic/webSearch_data/aircor2code.txt") 
    for line in file:
        row_data = line.strip("\n") 
        row_data_list = []
        split_data = row_data.split(" ")
        code = split_data[0]
        for i in range(1,len(split_data)):
            if split_data[i] != "":
                row_data_list.append(split_data[i])
        for item in row_data_list:
            if item in sent:
                file.close()
                return code
    file.close()
    return "00"

def check_airport_code(sent):#获得机场代码
    file = open(sys.path[0] + "/stastic/webSearch_data/airport2code.txt") 
    for line in file:
        row_data = line.strip("\n") 
        row_data_list = []
        split_data = row_data.split(" ")
        code = split_data[0]
        for i in range(1,len(split_data) - 1):
            if split_data[i] != "":
                row_data_list.append(split_data[i])
        for item in row_data_list:
            if item in sent:
                file.close()
                return code
    file.close()
    return "00"

def intent_recognization_func(sent):#将sent和所有语料做语义相似度匹配
    max_score = 0
    intent = "未识别到意图"
    max_score_list = []
    bert = get_bertvector()#获取Bert转句向量模型
    query_encode = bert.encode([sent])#将sent转成encode
    query2encode = get_query2encode()#将已有语料的encode读入
    score_dict = {}

    start = time.time()
    encode_list = []
    #new_array = np.zeros((len(query2encode),len(query_encode[0])))
    i = 0
    for key in query2encode:#所有语料的encode拼成矩阵
        encode_list.append(query2encode[key])
    matrix_a = np.mat(encode_list)#所有语料的encode拼成矩阵
    matrix_b = np.mat(query_encode[0])
    matrix_rsult = spatial.distance.cdist(matrix_a, matrix_b, 'cosine')#计算相似度
    #print(matrix_rsult)
    i = 0
    for key in query2encode:
        score_dict[key] = 1 - matrix_rsult[i].tolist()[0]
        i += 1
    end = time.time()
    print(str(round(end-start,5))+'s')

    query2intent = get_query2intent()
    intent2score = {}
    for key in score_dict:
        if key in query2intent:
            query_intent = query2intent[key]
            if query_intent not in intent2score:
                intent2score[query_intent] = [score_dict[key]]
            else:
                intent2score[query_intent].append(score_dict[key])
    for key in intent2score:
        if len(intent2score[key]) <= 5:
            intent2score[key] = float(sum(intent2score[key]) / len(intent2score[key]))
        else:
            top_five_score_list = heapq.nlargest(5,intent2score[key])
            intent2score[key] = float(sum(top_five_score_list) / 5)
    intent_max = max(intent2score, key=lambda x: intent2score[x])
    print(intent_max,intent2score[intent_max])
    return intent_max,intent2score[intent_max]
    
    # top_five_list = []
    # #a = sorted(score_dict.items(), key=lambda x: x[1], reverse=True) #根据相似度从大到小排序
    # for i in range(5):#选相似度最大的五个语料
    #     res_max = max(score_dict, key=lambda x: score_dict[x])
    #     top_five_list.append((res_max,score_dict[res_max]))
    #     score_dict.pop(res_max)
    # return top_five_list

def intent_recognization_from3w_func(sent):#从3W条语料中匹配最相近的五条语料
    max_score = 0
    intent = "未识别到意图"
    max_score_list = []
    bert = get_bertvector()
    query_encode = bert.encode([sent])
    query2encode = get_query2encode()
    score_dict = {}
    encode_list = []
    for key in query2encode:
        #score_dict[key] = cosine_similarity(query_encode[0],query2encode[key])
        encode_list.append(query2encode[key])
    matrix_a = np.mat(encode_list)
    matrix_b = np.mat(query_encode[0])
    matrix_rsult = spatial.distance.cdist(matrix_a, matrix_b, 'cosine')
    i = 0
    for key in query2encode:
        score_dict[key] = 1 - matrix_rsult[i].tolist()[0]
        i += 1
    
    query2intent = get_query2intent()
    intent2score = {}
    for key in score_dict:
        if key in query2intent:
            intent_list = query2intent[key]
            for query_intent in intent_list:
            #query_intent = query2intent[key]
            #query_item = query2intent[key]
            #query_intent = query_item["intent_level_1"] + "@" + query_item["intent_level_2"] + "@" + query_item["intent_level_3"]
                if query_intent not in intent2score:
                    intent2score[query_intent] = [score_dict[key]]
                else:
                    intent2score[query_intent].append(score_dict[key])
    for key in intent2score:
        if len(intent2score[key]) <= 5:
            intent2score[key] = float(sum(intent2score[key]) / len(intent2score[key]))
        else:
            top_five_score_list = heapq.nlargest(5,intent2score[key])
            intent2score[key] = float(sum(top_five_score_list) / 5)
    intent_max = max(intent2score, key=lambda x: intent2score[x])
    print(intent_max,intent2score[intent_max])
    
    intent_recognization_from3w_result = {"intent":[],"query":[],"high_score":[]}
    result_intent_list = []#去除结果中的重复意图
    if intent2score[intent_max] > 0.95:
        #top_five_list.append((intent_max,intent2score[intent_max]))
        intent_recognization_from3w_result["intent"].append((intent_max,intent2score[intent_max]))
        result_intent_list.append(intent_max)
        for j in range(2):
            res_max = max(score_dict, key=lambda x: score_dict[x])
            #print(res_max,score_dict[res_max])
            if res_max in query2intent:
                query_intent = query2intent[res_max]
            else:
                query_intent = res_max
            if query_intent in result_intent_list:
                score_dict.pop(res_max)
                continue
            if score_dict[res_max] > 0.99:
                intent_recognization_from3w_result["high_score"].append((res_max,score_dict[res_max]))
                result_intent_list.append(query_intent)
                score_dict.pop(res_max)
                continue
            if (check_aircor_code(sent) == check_aircor_code(res_max) and check_aircor_code(sent) != "00" )or (check_airport_code(sent) == check_airport_code(res_max) and check_airport_code(sent) != "00"):
                intent_recognization_from3w_result["query"].append((res_max,score_dict[res_max] + 0.1))
                result_intent_list.append(query_intent)
            else:
                intent_recognization_from3w_result["query"].append((res_max,score_dict[res_max]))
                result_intent_list.append(query_intent)
            #top_five_list.append((res_max,score_dict[res_max]))
            score_dict.pop(res_max)
    else:
        for j in range(5):
            res_max = max(score_dict, key=lambda x: score_dict[x])
            #print(res_max,score_dict[res_max])
            if res_max in query2intent:
                query_intent = query2intent[res_max]
            else:
                query_intent = res_max
            if query_intent in result_intent_list:
                score_dict.pop(res_max)
                continue
            if score_dict[res_max] > 0.99:
                intent_recognization_from3w_result["high_score"].append((res_max,score_dict[res_max]))
                result_intent_list.append(query_intent)
                score_dict.pop(res_max)
                continue
            if (check_aircor_code(sent) == check_aircor_code(res_max) and check_aircor_code(sent) != "00" )or (check_airport_code(sent) == check_airport_code(res_max) and check_airport_code(sent) != "00"):
                intent_recognization_from3w_result["query"].append((res_max,score_dict[res_max] + 0.1))
                result_intent_list.append(query_intent)
            else:
                intent_recognization_from3w_result["query"].append((res_max,score_dict[res_max]))
                result_intent_list.append(query_intent)
            #top_five_list.append((res_max,score_dict[res_max]))
            score_dict.pop(res_max)

    print(intent_recognization_from3w_result)
    return intent_recognization_from3w_result


def k_ford_intent_rec_test():#意图识别模型测试
    '''
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/final_data_1209.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    table_data = []
    for i in range(1,rows):
        row_data = table.row_values(i)
        table_data.append((row_data[0],row_data[1]))

    random.shuffle(table_data)

    tabel_lenth = len(table_data)
    for n in range(k):
        # 创建一个workbook 设置编码
        workbook = xlwt.Workbook(encoding='utf-8')
        # 创建一个worksheet
        worksheet = workbook.add_sheet('sheet 1')
        kk = 1
        #print(n * tabel_lenth // k,(n + 1) * tabel_lenth // k)
        for i in range(n * tabel_lenth // k,(n + 1) * tabel_lenth // k):
            row_data = table_data[i]
            for j in range(len(row_data)):
                worksheet.write(kk, j, label = row_data[j])
            kk += 1
        workbook.save('stastic/webSearch_data/cross-validation/train_data_' + str(n) + '.xlsx')
    '''
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/test-data_XGSS.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows

    # 创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')
    kk = 1
    for i in range(1000,rows):
        print(i)
        if i == 10000:
            break
        start = time.time()
        row_data = table.row_values(i)
        if row_data[1] == "":
            continue
        for j in range(4):
            worksheet.write(kk, j, label = row_data[j])
        try:
            sent = str(row_data[1])
            aircor = check_aircor(sent)
            aircor_dict = {'南方航空':'南航','海南航空':'海航','中国国际航空':'国航'}
            for key in aircor_dict:
                if key in sent:
                    sent = sent.replace(key,"")
                if aircor_dict[key] in sent:
                    sent = sent.replace(aircor_dict[key],"")
            result_data = web_search_func(sent,aircor)
            j = 4
            for row_result_data in result_data:
                worksheet.write(kk, j, label = str((row_result_data["title"],row_result_data["score"])))
                j += 1
                worksheet.write(kk, j, label = row_result_data["content"])
                j += 1

            kk += 1 #重排序
            j = 4
            sorted_data_dict = {}
            for row_result_data in result_data:
                if row_result_data["score"] > 0.98:
                    worksheet.write(kk, j, label = str((row_result_data["title"],row_result_data["score"])))
                    j += 1
                    worksheet.write(kk, j, label = row_result_data["content"])
                    j += 1
                else:
                    sorted_data_dict[row_result_data["title"]] = row_result_data["content"]
            #print(sorted_data_dict)
            sorted_key2score = {}
            for key in sorted_data_dict:
                sorted_url = "http://10.1.1.6:8180/api/count_sentence_similarity?sent1=" + sent + "&sent2=" + key
                r = requests.get(sorted_url)
                sorted_score = r.json()[0]
                sorted_key2score[key] = sorted_score
            sorted_key2score = sorted(sorted_key2score.items(), key=lambda item:item[1],reverse=True)#排序
            #print(sorted_key2score)
            for item in sorted_key2score:
                worksheet.write(kk, j, label = str(item))
                j += 1
                worksheet.write(kk, j, label = sorted_data_dict[item[0]])
                j += 1
        except Exception as e:
            #print(e)
            traceback.print_exc()
        kk += 1
        end = time.time()
        print(str(round(end-start,3))+'s')
    workbook.save('stastic/webSearch_data/result_data_sorted_01.xlsx')

    

def test(): #测试
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/CQA_data_test.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows

    #创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')
    k = 1
    query2encode = get_query2encode()
    for i in range(1,rows):
        #try:
        row_data = table.row_values(i)
        #start = time.time()
        if row_data[0] in query2encode:
            continue
        intent,intent_score = intent_recognization_func(row_data[0])
        #end = time.time()
        #print(str(round(end-start,3))+'s')
        worksheet.write(k, 0, label = row_data[0])
        rel_intent = ""
        if row_data[3] != "":
            rel_intent = row_data[3]
        else:
            rel_intent = row_data[2]
        worksheet.write(k, 1, label = rel_intent)
        worksheet.write(k, 2, label = intent)
        worksheet.write(k, 3, label = intent_score)
        if intent == rel_intent:
            worksheet.write(k, 4, label = 1)
        else:
            worksheet.write(k, 4, label = 0)
        k += 1
        # except Exception as e:
        #     print(e)
    workbook.save(sys.path[0] + '/stastic/webSearch_data/CQA_test_result.xlsx')

def delete_repeat(): #去重
    data = xlrd.open_workbook( 'new_data.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    query_list = []
    #创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')
    k = 1
    for i in range(1,rows):
        row_data = table.row_values(i)
        #start = time.time()
        if row_data[0] in query_list:
            continue
        else:
            for j in range(7):
                worksheet.write(k, j, label = row_data[j])
        k += 1
        query_list.append(row_data[0])
        # except Exception as e:
        #     print(e)
    workbook.save('new_data_1109.xlsx')

def merge_intent_data():#合并标注语料
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/百科_test_标注_丽颖.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    data_dict = {}
    for i in range(1,rows):
        row_data = table.row_values(i)
        if row_data[1] != "" and row_data[0] not in data_dict:
            data_dict[row_data[0]] = row_data[1]

    #创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')

    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/百科问答待标注语料_谭智隆.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    k = 1
    for i in range(1,rows):
        row_data = table.row_values(i)
        if i < 13424 or i > 19547:
            worksheet.write(k, 0, label = row_data[0])
            worksheet.write(k, 1, label = row_data[1])
            k += 1
        else:
            worksheet.write(k, 0, label = row_data[0])
            if row_data[0] in data_dict:
                worksheet.write(k, 1, label = data_dict[row_data[0]])
            k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/intent_data_1202.xlsx')

def annotate_data():#标注语料
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/百科问答意图体系1209_withanswer.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows

    #创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')
    k = 1

    for i in range(1,rows):
        row_data = table.row_values(i)
        if row_data[7] == "":
            for j in range(len(row_data)):
                worksheet.write(k, j, label = row_data[j])
            k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/待整理答案意图.xlsx')
    
def get_answer():
    intent2anwser = {}
    with open(sys.path[0] + '/stastic/webSearch_data/web2intent.json','r',encoding='utf8') as fp:
        web2intent_data = json.load(fp)
    fp.close()
    for aircor in web2intent_data:
        for web in web2intent_data[aircor]:
            for intent in web2intent_data[aircor][web]:
                if intent not in intent2anwser:
                    intent2anwser[intent] = [aircor + "@" + web]
                else:
                    intent2anwser[intent].append(aircor + "@" + web)

    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/CQA_answer.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1,rows):
        row_data = table.row_values(i)
        if row_data[2] != "":
            if row_data[1] not in intent2anwser:
                intent2anwser[row_data[1]] = [row_data[2]]
            else:
                intent2anwser[row_data[1]].append(row_data[2])

    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/airport_answer.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1,rows):
        row_data = table.row_values(i)
        if row_data[1] != "" and row_data[4] != "":
            if row_data[1] not in intent2anwser:
                intent2anwser[row_data[1]] = [row_data[4]]
            else:
                intent2anwser[row_data[1]].append(row_data[4])

    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/百科问答标注语料_1208.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    #创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')
    k = 1
    for i in range(1,rows):
        row_data = table.row_values(i)
        for j in range(0,7):
            worksheet.write(k, j, label = row_data[j])
        if row_data[2] != "":
            intent = row_data[2]
        else:
            intent = row_data[1]
        if intent in intent2anwser:
            print(intent2anwser[intent])
            for j in range(7,7 + len(intent2anwser[intent])):
                worksheet.write(k, j, label = intent2anwser[intent][j - 7])
        k += 1
        
    workbook.save(sys.path[0] + '/stastic/webSearch_data/百科问答标注语料_1209_with_answer.xlsx')

def get_answer_other():
    intent2anwser = {}
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/待整理答案意图.xls')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1,rows):
        row_data = table.row_values(i)
        intent = row_data[0]
        if row_data[1] != "":
            intent = row_data[1]
        if row_data[2] != "":
            intent = row_data[2]
        intent2anwser[intent] = row_data[7]
        
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/百科问答意图体系1209_withanswer.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    #创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')
    k = 1
    for i in range(1,rows):
        row_data = table.row_values(i)
        for j in range(0,len(row_data)):
            if row_data[j] != "":
                worksheet.write(k, j, label = row_data[j])
        intent = row_data[0]
        if row_data[1] != "":
            intent = row_data[1]
        if row_data[2] != "":
            intent = row_data[2]
        if intent in intent2anwser:
            print(intent2anwser[intent])
            worksheet.write(k, 7, label = intent2anwser[intent])
        k += 1
        
    workbook.save(sys.path[0] + '/stastic/webSearch_data/百科问答意图体系1214_withanswer.xlsx')



def data_7k_modify():#修正7k条语料的意图
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/百科问答意图体系1209_withanswer.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    intent_list = []
    for i in range(1,rows):
        row_data = table.row_values(i)
        if row_data[2] == "":
            intent = row_data[1]
        else:
            intent  = row_data[2]
        if intent not in intent_list:
            intent_list.append(intent)

    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/intent_data.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows  
    #创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')
    for i in range(1,rows):
        row_data = table.row_values(i)
        if row_data[3] == "":
            intent = row_data[2]
        else:
            intent  = row_data[3]
        worksheet.write(i, 0, label = row_data[0])
        if intent in intent_list:
            worksheet.write(i, 1, label = intent)

    workbook.save(sys.path[0] + '/stastic/webSearch_data/百科问答语料7k.xlsx')

def ume_data():#筛选航旅APP相关意图
    query2intent = {}
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/intent_data.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1,rows):
        row_data = table.row_values(i)
        intent = row_data[1]
        if row_data[2] != "":
            intent = row_data[2]
        if row_data[3] != "":
            intent = row_data[3]
        if row_data[0] not in query2intent:
            query2intent[row_data[0]] = intent

    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/final_data_7k.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    intent_list = []
    #创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建一个worksheet
    worksheet = workbook.add_sheet('sheet 1')
    for i in range(1,rows):
        row_data = table.row_values(i)
        for j in range(2):
                worksheet.write(i, j, label = row_data[j])
        if row_data[1] == "" and row_data[0] in query2intent:
            worksheet.write(i, 2, label = query2intent[row_data[0]])

    workbook.save(sys.path[0] + '/stastic/webSearch_data/待标注语料_7k.xlsx')

def get_fine_tune_data():
    query2intent = {}
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/final_data_1209.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1,rows):
        row_data = table.row_values(i)
        intent = row_data[1]
        if row_data[2] != "":
            intent = row_data[2]
        if row_data[3] != "":
            intent = row_data[3]
        if row_data[0] not in query2intent:
            query2intent[row_data[0]] = intent

def web_search_func(sent,aircor_name):#根据query返回页面
    intent_score = intent_recognization_from3w_func(sent)#从3W语料中选择最相关
    #intent_score,intent = intent_recognization_func(sent)#意图识别
    #result_score,doc,seg_list_without_stopword = str_match_func(sent,aircor_name)#字符串匹配
    
    doc = get_doc()
    stopword_list = []
    file = open(sys.path[0] + "/stastic/webSearch_data/stopwords-master/baidu_stopwords.txt") #加载停用词表
    for line in file:
        if line.strip("\n") not in stopword_list:
            stopword_list.append(line.strip("\n"))
    file.close()
    seg_list = jieba.lcut(sent,cut_all = False)#精确分词
    #seg_list2 = jieba.lcut_for_search(sent)
    seg_list_without_stopword = []#去停用词
    for word in seg_list:
        if word not in stopword_list:
            seg_list_without_stopword.append(word)

    final_result_list = []
    query2intent = get_query2intent()
    intent2answer = get_intent2answer()
    #final_result_tag_list = []
    query_list = []
    result_answer_list = []
    if intent_score["intent"] != []:#处理意图
        item = intent_score["intent"][0]
        intent = item[0]
        answer = "该问题答案暂缺，相关专家正在编辑..."
        web_title = "未找到相关页面"
        #source = intent2answer[intent]["source"]
        source = "暂缺"
        if intent in intent2answer:
            answer_list = intent2answer[intent]["answer"]
            if answer_list == []:#没答案，则答案为意图
                answer = "一级意图:" + intent2answer[intent]["intent_level_1"] 
                if intent2answer[intent]["intent_level_2"] != "":
                    answer = answer + "，二级意图:" + intent2answer[intent]["intent_level_2"] 
                if intent2answer[intent]["intent_level_3"] != "":
                    answer = answer + "，三级意图:" + intent2answer[intent]["intent_level_3"] 
                url_str = "/api/content_search?query=" + web_title + "&answer=" + answer
                description,high_line = answer,[]
                #description_str = ",description:" + description + "high_line:" + str(high_line)
                final_result_list.append({"title":intent.strip("@"),"href":url_str,"web_part":2,"source":source,"score":item[1],"content":description + "...","high_line":high_line})
            else:
                result_flag = 0 
                for answer_item in answer_list:
                    if "@" not in answer_item:
                        answer = answer_item
                        result_flag = 1
                    else:
                        if answer_item.split("@")[0] == aircor_name or aircor_name == "所有航司":
                            answer = answer_item
                            web_title = answer_item.split("@")[1]
                            result_flag = 1
                    if answer == "该问题答案暂缺，相关专家正在编辑..." :
                        continue
                    if answer in result_answer_list:
                        continue
                    else:
                        result_answer_list.append(answer)
                    if web_title != "未找到相关页面":
                        with open(sys.path[0] + '/stastic/webSearch_data/all_data.json','r',encoding='utf8') as fp:
                            json_data = json.load(fp)
                        fp.close()
                        # if aircor_name == "所有航司":
                        #     aircor_list = ["海南航空" , "南方航空", "中国国际航空"]
                        # elif aircor_name == "其他航司":
                        #     aircor_list = []
                        # else:
                        #     aircor_list = [aircor_name]
                        if answer_item.split("@")[0] in ["海南航空" , "南方航空", "中国国际航空"]:
                            aircor_list = [answer_item.split("@")[0]]
                        else:
                            aircor_list = []
                        for aircor in aircor_list:
                            for kk in json_data[aircor]:
                                for key in json_data[aircor][kk]:
                                    for sub_dict in json_data[aircor][kk][key]:
                                        for sub_key in sub_dict:
                                            if sub_key == web_title:
                                                web_title = aircor + "@" + kk + "@" + key + "@" + sub_key
                                                answer = ""
                                                source = aircor
                                                break
                    url_str = "/api/content_search?query=" + web_title + "&answer=" + answer
                    if web_title in doc:
                        item_text = "".join(doc[web_title])
                        item_text_list = []
                        for t in item_text.split("。"):
                            item_text_list.append(t.strip("，"))
                        description,high_line = get_query_description(seg_list_without_stopword,item_text_list)
                    else:
                        description,high_line = answer,[]
                    description_str = ",description:" + description + "high_line:" + str(high_line)
                    final_result_list.append({"title":intent.strip("@"),"href":url_str,"web_part":2,"source":source,"score":item[1],"content":description + "...","high_line":high_line})
                if result_flag == 0 :
                    answer = "一级意图:" + intent2answer[intent]["intent_level_1"] 
                    if intent2answer[intent]["intent_level_2"] != "":
                        answer = answer + "，二级意图:" + intent2answer[intent]["intent_level_2"] 
                    if intent2answer[intent]["intent_level_3"] != "":
                        answer = answer + "，三级意图:" + intent2answer[intent]["intent_level_3"] 
                    url_str = "/api/content_search?query=" + web_title + "&answer=" + answer
                    description,high_line = answer,[]
                    #description_str = ",description:" + description + "high_line:" + str(high_line)
                    final_result_list.append({"title":intent.strip("@"),"href":url_str,"web_part":2,"source":source,"score":item[1],"content":description + "...","high_line":high_line})
        
    for item in intent_score["query"]:#处理query
        query = item[0]
        if query in query2intent:
            intent_list = query2intent[query]
            for query_intent in intent_list:
            #query_intent = query2intent[query]#意图
                if query_intent in intent2answer:
                    answer_list = intent2answer[query_intent]["answer"]
                    if answer_list == []:#没答案，则答案为意图
                        web_title = "未找到相关页面"
                        answer = "一级意图:" + intent2answer[query_intent]["intent_level_1"] 
                        if intent2answer[query_intent]["intent_level_2"] != "":
                            answer = answer + "，二级意图:" + intent2answer[query_intent]["intent_level_2"] 
                        if intent2answer[query_intent]["intent_level_3"] != "":
                            answer = answer + "，三级意图:" + intent2answer[query_intent]["intent_level_3"] 
                        url_str = "/api/content_search?query=" + web_title + "&answer=" + answer
                        description,high_line = answer,[]
                        source = intent2answer[query_intent]["source"]
                        #description_str = ",description:" + description + "high_line:" + str(high_line)
                        final_result_list.append({"title":query.strip("@"),"href":url_str,"web_part":2,"source":source,"score":item[1],"content":description + "...","high_line":high_line})
                    else:
                        result_flag = 0 
                        for answer_item in answer_list:
                            answer = "该问题答案暂缺，相关专家正在编辑..."
                            web_title = "未找到相关页面"
                            source = "暂缺"
                            if "@" not in answer_item:
                                answer = answer_item
                                result_flag = 1
                            else:
                                if answer_item.split("@")[0] == aircor_name or aircor_name == "所有航司":
                                    answer = answer_item
                                    web_title = answer_item.split("@")[1]
                                    result_flag = 1
                            if answer == "该问题答案暂缺，相关专家正在编辑...":
                                continue
                            if answer in result_answer_list:
                                continue
                            else:
                                result_answer_list.append(answer)
                            if web_title != "未找到相关页面":
                                with open(sys.path[0] + '/stastic/webSearch_data/all_data.json','r',encoding='utf8') as fp:
                                    json_data = json.load(fp)
                                fp.close()
                                # if aircor_name == "所有航司":
                                #     aircor_list = ["海南航空" , "南方航空", "中国国际航空"]
                                # elif aircor_name == "其他航司":
                                #     aircor_list = []
                                # else:
                                #     aircor_list = [aircor_name]
                                if answer_item.split("@")[0] in ["海南航空" , "南方航空", "中国国际航空"]:
                                    aircor_list = [answer_item.split("@")[0]]
                                else:
                                    aircor_list = []
                                for aircor in aircor_list:
                                    for kk in json_data[aircor]:
                                        for key in json_data[aircor][kk]:
                                            for sub_dict in json_data[aircor][kk][key]:
                                                for sub_key in sub_dict:
                                                    if sub_key == web_title:
                                                        web_title = aircor + "@" + kk + "@" + key + "@" + sub_key
                                                        answer = ""
                                                        #source = web_title.split("@")[0]
                                                        source = aircor
                                                        break
                            url_str = "/api/content_search?query=" + web_title + "&answer=" + answer
                            if web_title in doc:
                                item_text = "".join(doc[web_title])
                                item_text_list = []
                                for t in item_text.split("。"):
                                    item_text_list.append(t.strip("，"))
                                description,high_line = get_query_description(seg_list_without_stopword,item_text_list)
                            else:
                                description,high_line = answer,[]
                            description_str = ",description:" + description + "high_line:" + str(high_line)
                            final_result_list.append({"title":query.strip("@"),"href":url_str,"web_part":2,"source":source,"score":item[1],"content":description + "...","high_line":high_line})
                        if result_flag == 0 :
                            answer = "一级意图:" + intent2answer[query_intent]["intent_level_1"] 
                            if intent2answer[query_intent]["intent_level_2"] != "":
                                answer = answer + "，二级意图:" + intent2answer[query_intent]["intent_level_2"] 
                            if intent2answer[query_intent]["intent_level_3"] != "":
                                answer = answer + "，三级意图:" + intent2answer[query_intent]["intent_level_3"] 
                            url_str = "/api/content_search?query=" + web_title + "&answer=" + answer
                            description,high_line = answer,[]
                            #description_str = ",description:" + description + "high_line:" + str(high_line)
                            final_result_list.append({"title":query.strip("@"),"href":url_str,"web_part":2,"source":source,"score":item[1],"content":description + "...","high_line":high_line})
        else:
            answer = "该问题答案暂缺，相关专家正在编辑..."
            web_title = "未找到相关页面"
            url_str = "/api/content_search?query=" + web_title + "&answer=" + answer
            source = "暂缺"
            description,high_line = answer,[]
            final_result_list.append({"title":query.strip("@"),"href":url_str,"web_part":2,"source":source,"score":item[1],"content":description + "...","high_line":high_line})

    return final_result_list
   
def sorted_data_process():
    query2intent = get_query2intent()
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/sorted_data.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    data_list = []
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    k = 1
    for i in range(1,rows):
        print(i)
        row_data = table.row_values(i)
        sub_data = row_data[0]
        for j in range(1,len(row_data)):
            if row_data[j] != "":
                obj_data = row_data[j]
                if set([sub_data,obj_data]) in data_list:
                    continue
                data_list.append(set([sub_data,obj_data]))
                if sub_data == obj_data or obj_data not in query2intent:
                    continue
                if obj_data in query2intent and len(query2intent[obj_data]) == 0:
                    continue
                worksheet.cell(k, 1, value = sub_data)
                worksheet.cell(k, 2, value = obj_data)
                x = [val for val in query2intent[sub_data] if val in query2intent[obj_data]]
                if len(x) > 0:
                    worksheet.cell(k, 3, value = 1)
                else:
                    worksheet.cell(k, 3, value = 0)
                k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/sorted_data/sorted_data_train.xlsx')
    
def sorted_data_process_2():
    query2intent = get_query2intent()
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/sorted_data/sorted_data_train.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    data_list = []
    for i in range(0,rows):
        row_data = table.row_values(i)
        data_list.append(set([row_data[0],row_data[1]]))

    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/final_data_4w2.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    table_data = []
    for i in range(1,rows):
        row_data = table.row_values(i)
        table_data.append(row_data)

    k = 1
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    while k <= 100000:
        print(k)
        x = np.random.randint(0,len(table_data) - 1,1)[0]
        y = np.random.randint(0,len(table_data) - 1,1)[0]
        #print(x,y,table_data[x],table_data[y])
        if table_data[x][1] == "" and table_data[y][1] == "":
            continue
        worksheet.cell(k, 1, value = table_data[x][0])
        worksheet.cell(k, 2, value = table_data[y][0])
        if table_data[x][0] in query2intent:
            x_intent = query2intent[table_data[x][0]]
        else:
            x_intent = []
        if table_data[y][0] in query2intent:
            y_intent = query2intent[table_data[y][0]]
        else:
            y_intent = []
        z = [val for val in x_intent if val in y_intent]
        if len(z) > 0:
            worksheet.cell(k, 3, value = 1)
        else:
            worksheet.cell(k, 3, value = 0)
        k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/sorted_data/sorted_data_train_2.xlsx')

def sorted_data_process_3():
    data_list = []
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/sorted_data/sorted_data_train.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(0,rows):
        row_data = table.row_values(i)
        data_list.append(row_data)

    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/sorted_data/sorted_data_train_2.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(0,rows):
        row_data = table.row_values(i)
        data_list.append(row_data)
    print(len(data_list))
    
    random.shuffle(data_list)
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    k = 1
    for i in range(0, int(len(data_list) * 0.8)):
        worksheet.cell(k, 1, value = data_list[i][0])
        worksheet.cell(k, 2, value = data_list[i][1])
        worksheet.cell(k, 3, value = data_list[i][2])
        k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/sorted_data/train.xlsx')

    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    k = 1
    for i in range(int(len(data_list) * 0.8),int(len(data_list) * 0.9)):
        worksheet.cell(k, 1, value = data_list[i][0])
        worksheet.cell(k, 2, value = data_list[i][1])
        worksheet.cell(k, 3, value = data_list[i][2])
        k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/sorted_data/test.xlsx')

    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    k = 1
    for i in range(int(len(data_list) * 0.9),len(data_list)):
        worksheet.cell(k, 1, value = data_list[i][0])
        worksheet.cell(k, 2, value = data_list[i][1])
        worksheet.cell(k, 3, value = data_list[i][2])
        k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/sorted_data/dev.xlsx')

def get_mysqlSpider_data():
    db = mysql.connect("172.16.101.209", "spider", "spider_123", "spider",port=3306,charset='utf8')
    cursor = db.cursor()
    sql = "SELECT * FROM umetrip_web_content_crawl limit 100,300"
    #sql = "show columns FROM umetrip_web_content_crawl"
    cursor.execute(sql)
    results = cursor.fetchall() #获取全部结果集。 
    db.commit()
    if not results: #判断是否为空。
        print("数据为空！")
    else:
        results_list = []
        for row_data in results:
            #print(row_data)
            row_data_dict = {}
            row_data_dict["title"] = row_data[3]
            row_data_dict["content"] = row_data[6]
            results_list.append(row_data_dict)
    with open(sys.path[0] + "/stastic/webSearch_data/umeSpider_data_exampel.json", "w", encoding='utf-8') as fp:
        fp.write(json.dumps(results_list, ensure_ascii=False, indent=4))

database_address = "http://10.5.150.11:7474"
import requests, json

def execute(statement):
    url = database_address + "/db/data/transaction/commit"
    #auth = 'Basic ' + (base64.b64encode((config.db_username + ':' + config.db_password).encode('utf-8'))).decode('utf-8')
    header = {"Accept": "application/json; charset=UTF-8",
              "Content-Type": "application/json",
              "Authorization": "Basic bmVvNGo6MXFhejJ3c3g="}

    payload = {
        "statements": [{
            "statement": statement
        }]
    }
    #urllib3.PoolManager(num_pools=10000)
    print("查询" + database_address + "数据库: " + statement)
    #print(requests.post(url, headers=header, json=payload))
    response = requests.post(url, headers=header, json=payload).json()
    #print(response)
    if len(response["results"]) != 0:
        return response["results"][0]["data"]
    else:
        return []

def neo4j_data():
    statement = "match (n) where n.名称 = \"味千拉面\" return id(n),n"
    result = execute(statement)
    for item in result:
        print(item)

def get_data_from_other():#从机场问答和小横问答补充语料
    data_list = []
    airport_data = {}
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/datasets_1201_airport.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(0,rows):
        row_data = table.row_values(i)
        if row_data[0] == "":
            continue
        if row_data[1] not in airport_data:
            airport_data[row_data[1]] = [row_data[0]]
        else:
            airport_data[row_data[1]].append(row_data[0])

    xiaoheng_data = {}
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/datasets_0814_小横问答.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(0,rows):
        row_data = table.row_values(i)
        if row_data[0] == "":
            continue
        if row_data[1] not in xiaoheng_data:
            xiaoheng_data[row_data[1]] = [row_data[0]]
        else:
            xiaoheng_data[row_data[1]].append(row_data[0])

    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)

    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/百科问答意图语料数量统计1221_合.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    k = 1
    for i in range(0,rows):
        row_data = table.row_values(i)
        airport_intent = row_data[4]
        xiaoheng_intent = row_data[5]
        if airport_intent != "":
            if airport_intent in airport_data:
                for item in airport_data[airport_intent]:
                    if item in data_list:
                        continue
                    worksheet.cell(k, 1, value = item)
                    worksheet.cell(k, 2, value = row_data[0])
                    if row_data[8] == "重新标注":
                        worksheet.cell(k, 3, value = "重新标注")
                    data_list.append(item)
                    k += 1
        if xiaoheng_intent != "":
            if xiaoheng_intent in xiaoheng_data:
                for item in xiaoheng_data[xiaoheng_intent]:
                    if item in data_list:
                        continue
                    worksheet.cell(k, 1, value = item)
                    worksheet.cell(k, 2, value = row_data[0])
                    if row_data[8] == "重新标注":
                        worksheet.cell(k, 3, value = "重新标注")
                    data_list.append(item)
                    k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/new_data_1228.xlsx')


def check_airport_code_and_name(sent):#获得机场code和名称
    file = open(sys.path[0] + "/stastic/webSearch_data/airport2code.txt") 
    airport_name = "未识别到机场名称"
    code = "00"
    for line in file:
        row_data = line.strip("\n") 
        row_data_list = []
        split_data = row_data.split(" ")
        for i in range(1,len(split_data) - 1):
            if split_data[i] != "":
                row_data_list.append(split_data[i])
        for item in row_data_list:
            if item in sent:
                if airport_name == "未识别到机场名称":
                    airport_name = item
                    code = split_data[0]
                else:
                    if len(airport_name) < len(item):
                        airport_name = item
                        code = split_data[0]
    file.close()
    return code,airport_name

def check_aircor_code_and_name(sent):#获得航司code和名称
    file = open(sys.path[0] + "/stastic/webSearch_data/aircor2code.txt") 
    code = "00"
    aircor_name = "未识别到航司名称"
    for line in file:
        row_data = line.strip("\n") 
        row_data_list = []
        split_data = row_data.split(" ")
        for i in range(1,len(split_data)):
            if split_data[i] != "":
                row_data_list.append(split_data[i])
        for item in row_data_list:
            if item in sent:
                if aircor_name == "未识别到航司名称":
                    aircor_name = item
                    code = split_data[0]
                else:
                    if len(aircor_name) < len(item):
                        aircor_name = item
                        code = split_data[0]
    file.close()
    return code,aircor_name

def check_trip_number(sent):#将符合航班号格式的字符串替换为航班号
    #cop_1 = re.compile("(?<![A-Za-z])([A-Za-z]{2}|\d[A-Za-z]|[A-Za-z]\d)\d{3,4}(?!\d)") 
    cop_1 = re.compile("(?<![\dA-Z])(?!\d{2})([A-Z\d]{2})\s?(\d{2,4})(?!\d)",re.I) 
    if cop_1.search(sent) is not None:
        a = cop_1.search(sent).span()
        #print(a)
        trip_number = sent[a[0]:a[1]]
    else:
        trip_number = ""
    string = cop_1.sub('航班号', sent) 
    return trip_number,string

def check_city_name(sent):#将城市名进行替换
    file = open(sys.path[0] + "/stastic/webSearch_data/city_set.txt") 
    for line in file:
        city_name = line.strip("\n").strip("市") 
        if city_name in sent:
            return city_name
    return "未识别到城市名称"

def replace_entity2label(sent):#替换句中的航司、机场、城市名、航班号
    airport_name_list,aircor_name_list,city_name_list = [],[],[]
    airport_code,airport_name = check_airport_code_and_name(sent)
    if airport_code != "00":
        sent = sent.replace(airport_name,"机场")
        airport_name_list.append(airport_name)
        airport_code,airport_name = check_airport_code_and_name(sent)
        if airport_code != "00":
            sent = sent.replace(airport_name,"机场")
            airport_name_list.append(airport_name)
    aircor_code,aircor_name = check_aircor_code_and_name(sent)
    if aircor_code != "00":
        sent = sent.replace(aircor_name,"航空公司")
        aircor_name_list.append(aircor_name)
        aircor_code,aircor_name = check_aircor_code_and_name(sent)
        if aircor_code != "00":
            sent = sent.replace(aircor_name,"航空公司")
            aircor_name_list.append(aircor_name)
    city_name = check_city_name(sent)
    if city_name != "未识别到城市名称":
        sent = sent.replace(city_name,"城市名")
        city_name_list.append(city_name)
        city_name = check_city_name(sent)
        if city_name != "未识别到城市名称":
            sent = sent.replace(city_name,"城市名")
            city_name_list.append(city_name)
    trip_number,sent = check_trip_number(sent)
    return sent,airport_name_list,aircor_name_list,city_name_list,trip_number


def data_process_4w5():
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/query_data_4w5_processed.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows

    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    for i in range(1,rows):
        print(i)
        row_data = table.row_values(i)
        worksheet.cell(i + 1, 1, value = row_data[0])
        worksheet.cell(i + 1, 2, value = row_data[1])
    
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/query_data_4w5.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1,rows):
        print(i)
        row_data = table.row_values(i)
        worksheet.cell(i + 1, 3, value = row_data[0])
        worksheet.cell(i + 1, 4, value = replace_entity2label(row_data[0]))
    workbook.save(sys.path[0] + '/stastic/webSearch_data/query_data_4w5_processed_new.xlsx')


import csv
def xlsx2csv():
    all_label_list = ["|"]
    all_query_list = []
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/query_data_4w5_processed.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1,rows):
        row_data = table.row_values(i)
        query = row_data[0]
        if row_data[1] == "":
            continue
        label_list = row_data[1].strip("").split(";")
        for label in label_list:
            if label not in all_label_list and label != "":
                all_label_list.append(label)
    print(len(all_label_list))
    f = open('vocabulary_label.txt','w')
    for label in all_label_list:
        f.write(label + '\n')
    f.close()

    query2list_dict = {}
    for i in range(1,rows):
        row_data = table.row_values(i)
        query = row_data[0]
        query_label_onehot = [0] * len(all_label_list)
        if row_data[1] == "":
            query_label_onehot[0] = 1
        else:
            label_list = row_data[1].split(";")
            for label in label_list:
                if label != "":
                    query_label_onehot[all_label_list.index(label)] = 1
        if query not in all_query_list:
            all_query_list.append(query)
            query2list_dict[query] = query_label_onehot
    
    random.shuffle(all_query_list)
    with open("train_onehot_1.csv","w") as csvfile: 
        writer = csv.writer(csvfile)
        writer.writerow(["content"] + all_label_list)
        for query in all_query_list[:int(len(all_query_list) * 0.9)]:
            writer.writerow([query] + query2list_dict[query])
    csvfile.close()
    with open("test_onehot_1.csv","w") as csvfile: 
        writer = csv.writer(csvfile)
        writer.writerow(["content"] + all_label_list)
        for query in all_query_list[int(len(all_query_list) * 0.9):]:
            writer.writerow([query] + query2list_dict[query])
    csvfile.close()

def url_test():
    answer_url = "http://119.254.233.179/UMEKnowledgeGraphQA/restapi/qa_airport?content=%E5%AE%89%E6%A3%80%E6%B5%81%E7%A8%8B&airport=URC&callback"
    r = requests.get(answer_url)
    print(json.loads(r.text[1:-1])["parameter"][1])

def check_error_data():#找出标注错误的语料
    intent_list = []
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/schema_withanswer_1211.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1,rows):
        row_data = table.row_values(i)
        if "KBQA" in row_data[6]:
            continue
        if row_data[0] != "" :
            intent = row_data[0]
        if row_data[1] != "" :
            intent = row_data[1]
        if row_data[2] != "" :
            intent = row_data[2]
        if intent not in intent_list:
            intent_list.append(intent)
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)

    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/query_data_4w5_processed.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    k = 1
    for i in range(1,rows):
        row_data = table.row_values(i)
        item_intent_list = row_data[1].split(";")
        for intent in item_intent_list:
            if intent != "" and intent not in intent_list:
                worksheet.cell(k, 1, value = i + 1)
                worksheet.cell(k, 2, value = row_data[0])
                worksheet.cell(k, 3, value = row_data[1])
                k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/error_data.xlsx')

def qa_acc():#准确率统计
    with open(sys.path[0] + '/stastic/webSearch_data/test_result_0218.json','r',encoding='utf8') as fp:
        json_data = json.load(fp)
    fp.close()
    qa_num = 0
    top_1 = 0
    top_3 = 0
    top_5 = 0
    for item in json_data:
        qa_num += 1
        k = 0
        for answer_item in item["answers"]:
            k += 1
            if answer_item["atti"] == 1:
                top_5 += 1
                break
        if k <= 1:
            top_1 += 1 
        if k <= 3:
            top_3 += 1 
    print(top_1,top_3,top_5)
    print("总问题数:" + str(qa_num) + ",top_1准确率:" + str(round(top_1/qa_num,2)) + ",top_3准确率:" + str(round(top_3/qa_num,2)) + ",top_5准确率" + str(round(top_3/qa_num,2)) )
import ahttp 

def ahttp_test():
    sent = "南航值机柜台在哪里"
    # urls = [ f"https://movie.douban.com/top250?start={i*25}" for i in range(10) ]
    # reqs = [ahttp.get(url) for url in urls]
    # resps = ahttp.run(reqs)

    intent_score_url = "http://127.0.0.1:8180/api/intent_recognization_from3w?sent=" + sent #从intent_score的模型中召回
    fine_tuning_score_url = "http://127.0.0.1:8190/api/recall_query_by_bert?sent=" + sent #从fine-tuning后的模型中召回
    bm25_score_url = "http://127.0.0.1:8170/api/gensim_bm25?sent=" + sent #通过BM25召回
    classifier_multi_label_url = "http://127.0.0.1:8170/api/classifier_multi_label?sent=" + sent #通过textcnn召回
    urls = [intent_score_url,fine_tuning_score_url,bm25_score_url,classifier_multi_label_url]
    reqs = [ahttp.get(url) for url in urls]
    resps = ahttp.run(reqs, order=True)
    intent_score,fine_tuning_score,bm25_score,classifier_multi_label_list = resps[0].json(),resps[1].json(),resps[2].json(),resps[3].json()

    print(intent_score,fine_tuning_score,bm25_score,classifier_multi_label_list)
    return intent_score,fine_tuning_score,bm25_score,classifier_multi_label_list

def bdzd_xgss_data_process():#处理百度知道和相关搜索的数据
    data_list = []
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/test-data_XGSS.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1,rows):
        print(i)
        row_data = table.row_values(i)
        if row_data[1] not in data_list:
            data_list.append(row_data[1])
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/test-data_BDZD.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1,rows):
        row_data = table.row_values(i)
        if row_data[1] not in data_list:
            data_list.append(row_data[1])
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    k = 1
    for item in data_list:
        print(k)
        worksheet.cell(k, 1, value = item)
        k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/bdzd_and_xgss_data.xlsx')

def test_bdzd_xgss_data():
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/bdzd_and_xgss_data.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows

    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    for i in range(42800,rows):
        if i % 100 ==0 :
            print(i)
        row_data = table.row_values(i)
        sent = row_data[0]
        answer_url = "http://127.0.0.1:8180/api/web_search?query=" + str(sent) 
        r = requests.get(answer_url)
        try:
            result = r.json()
            #print(result)
            worksheet.cell(i + 1, 1, value = sent)
            worksheet.cell(i + 1, 2, value = result[0]["title"])
        except Exception as e:
            print(sent)
            traceback.print_exc()
        if i % 1000 == 0:
            workbook.save(sys.path[0] + '/stastic/webSearch_data/bdzd_xgss_test_result_0304.xlsx')
    workbook.save(sys.path[0] + '/stastic/webSearch_data/bdzd_xgss_test_result_0304.xlsx')

def answer_test():
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/answer_test_0315.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    k = 1
    for i in range(1,rows):
        row_data = table.row_values(i)
        print(i)
        if row_data[0] == "":
            continue
        for j in range(1,len(row_data) + 1):
            worksheet.cell(k, j, value = row_data[j - 1])
        sent = row_data[1]
        answer_url = "http://127.0.0.1:8180/api/web_search?query=" + str(sent) 
        r = requests.get(answer_url)
        try:
            result = r.json()
            #print(result)
            worksheet.cell(k, 12, value = result[0]["title"])
            worksheet.cell(k, 13, value = result[0]["source"])
        except Exception as e:
            print(sent)
            traceback.print_exc()
        k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/answer_test_0315_result.xlsx')

def intent_processed():#处理意图
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/answer_test_0315_result.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    for i in range(1,rows):
        row_data = table.row_values(i)
        for j in range(1,len(row_data) + 1):
            if j == 13 and row_data[j - 1] != "":
                intent = row_data[j - 1].split("--")[0]
                worksheet.cell(i, j, value = intent)
            else:
                worksheet.cell(i, j, value = row_data[j - 1])
    workbook.save(sys.path[0] + '/stastic/webSearch_data/answer_test_0315_result_new.xlsx')

def cluster_intent_data_processed():
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/clustering outcome bert.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    intent_now = ""
    query_dict = {}
    k = 1
    for i in range(1,rows):
        print(i)
        row_data = table.row_values(i)
        intent = row_data[1]
        query = row_data[2]
        label = row_data[4]
        if intent_now == "":
            intent_now = intent 
        if intent_now == intent:
            if label not in query_dict:
                query_dict[label] = [query]
            else:
                query_dict[label].append(query)
        else:
            for label in query_dict:
                for item in query_dict[label]:
                    worksheet.cell(k, 1, value = intent_now)
                    worksheet.cell(k, 2, value = item)
                    worksheet.cell(k, 3, value = label)
                    k += 1
            query_dict = {}
            query_dict[label] = [query]
            intent_now = intent
        if i == rows - 1:
            for label in query_dict:
                for item in query_dict[label]:
                    worksheet.cell(k, 1, value = intent_now)
                    worksheet.cell(k, 2, value = item)
                    worksheet.cell(k, 3, value = label)
                    k += 1
    workbook.save(sys.path[0] + '/stastic/webSearch_data/clustering_outcome_bert_new.xlsx')

def transformer_model(input_tensor,
                      attention_mask=None,
                      hidden_size=768,
                      num_hidden_layers=12,
                      num_hidden_groups=12,
                      num_attention_heads=12,
                      intermediate_size=3072,
                      inner_group_num=1,
                      intermediate_act_fn="gelu",
                      hidden_dropout_prob=0.1,
                      attention_probs_dropout_prob=0.1,
                      initializer_range=0.02,
                      do_return_all_layers=False):

  if hidden_size % num_attention_heads != 0:
    raise ValueError(
        "The hidden size (%d) is not a multiple of the number of attention "
        "heads (%d)" % (hidden_size, num_attention_heads))

  attention_head_size = hidden_size // num_attention_heads
  input_shape = get_shape_list(input_tensor, expected_rank=3)
  input_width = input_shape[2]

  all_layer_outputs = []
  if input_width != hidden_size:
    prev_output = dense_layer_2d(
        input_tensor, hidden_size, create_initializer(initializer_range),
        None, name="embedding_hidden_mapping_in")
  else:
    prev_output = input_tensor
  with tf.variable_scope("transformer", reuse=tf.AUTO_REUSE):
    for layer_idx in range(num_hidden_layers):
      group_idx = int(layer_idx / num_hidden_layers * num_hidden_groups)
      with tf.variable_scope("group_%d" % group_idx):
        with tf.name_scope("layer_%d" % layer_idx):
          layer_output = prev_output
          for inner_group_idx in range(inner_group_num):
            with tf.variable_scope("inner_group_%d" % inner_group_idx):
              layer_output = attention_ffn_block(
                  layer_output, hidden_size, attention_mask,
                  num_attention_heads, attention_head_size,
                  attention_probs_dropout_prob, intermediate_size,
                  intermediate_act_fn, initializer_range, hidden_dropout_prob)
              prev_output = layer_output
              all_layer_outputs.append(layer_output)
  if do_return_all_layers:
    return all_layer_outputs
  else:
    return all_layer_outputs[-1]

labelsent2intent = {}
def get_labelsent2intent():
    global labelsent2intent
    if labelsent2intent == {}:
        data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/query_data_4w5_processed_new_0322.xlsx')
        table = data.sheet_by_index(0)
        rows = table.nrows
        for i in range(1,rows):
            row_data = table.row_values(i)
            if row_data[0] not in labelsent2intent and row_data[1] != "":
                labelsent2intent[row_data[0]] = row_data[1]
    return labelsent2intent

def bdzd_data_add_intent():
    labelsent2intent = get_labelsent2intent()
    
    data = xlrd.open_workbook(sys.path[0] + '/stastic/webSearch_data/百度知道语料测试_待标注.xlsx')
    table = data.sheet_by_index(0)
    rows = table.nrows
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    for i in range(1,rows):
        if i % 100 == 0:
            print(i)
        row_data = table.row_values(i)
        worksheet.cell(i + 1, 1, value = row_data[0])
        worksheet.cell(i + 1, 2, value = row_data[1])
        worksheet.cell(i + 1, 5, value = row_data[3])
        sent = row_data[1]
        new_sent = ""
        sent,airport_name_list,aircor_name_list,city_name_list,trip_number = replace_entity2label(sent)
        if "T1" in sent:
            new_sent = sent.replace("T1","t1")
        if sent in labelsent2intent:
            intent = labelsent2intent[sent]
        elif new_sent in labelsent2intent:
            intent = labelsent2intent[new_sent]
        else:
            intent = ""
        worksheet.cell(i + 1, 3, value = intent)
        if i % 10000 == 0:
            workbook.save(sys.path[0] + '/stastic/webSearch_data/百度知道语料测试_待标注_加意图.xlsx')
    workbook.save(sys.path[0] + '/stastic/webSearch_data/百度知道语料测试_待标注_加意图.xlsx')

bdzd_data_add_intent()
#cluster_intent_data_processed()
#intent_processed()
#test_bdzd_xgss_data()
#bdzd_xgss_data_process()
#qa_acc()
#ahttp_test()
#check_trip_number("CA1515航班动态")
#data_process_4w5()
#def get_train_and_test_csv():
#neo4j_data()
#check_error_data()
#url_test()
#xlsx2csv()  
#data_process_4w5()
#intent_recognization_func()
#delete_repeat()
#test()
#annotate_data()
#get_answer()
#get_query2encode()
#data_7k_modify()
#get_answer_other()
#get_data_from_other()
#k_ford_intent_rec_test()
#sorted_data_process_3()
#get_mysqlSpider_data()
#sorted_data_process_2()
#get_query2encode()
#ume_data()
#merge_intent_data()
#data_process()